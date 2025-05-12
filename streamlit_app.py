import streamlit as st
import pandas as pd
from datetime import datetime
import os
import zipfile
from openpyxl import Workbook, load_workbook

# Path to Excel file
DATA_FILE = "all_branch_data.xlsx"

# ✅ Step 1: Ensure Excel file is valid (create or repair if needed)
def ensure_valid_excel(file_path):
    if not os.path.exists(file_path) or os.path.getsize(file_path) == 0:
        wb = Workbook()
        wb.save(file_path)
    else:
        try:
            with zipfile.ZipFile(file_path, 'r') as zf:
                zf.testzip()
        except zipfile.BadZipFile:
            wb = Workbook()
            wb.save(file_path)

# Run validation check
ensure_valid_excel(DATA_FILE)

# ✅ Branch Code → (Branch Name, [Riders])
branch_data = {
    "6661": ("Ali Branch", ["Rider A", "Rider B", "Rider C"]),
    "0001": ("Karachi Branch", ["Rider X", "Rider Y"]),
    "7860": ("Lahore Branch", ["Rider M", "Rider N", "Rider Z"]),
}

st.title("📦 Rider Slip Submission Form")

# Step 2: Branch code input
branch_code = st.text_input("Enter Your Branch Code")

if branch_code in branch_data:
    branch_name, riders = branch_data[branch_code]
    st.success(f"✅ Welcome, {branch_name}!")

    date = st.date_input("Select Date", datetime.today())
    rider = st.selectbox("Select Rider", riders)
    cash_slips = st.number_input("Cash Slips", min_value=0, step=1)
    online_slips = st.number_input("Online Slips", min_value=0, step=1)

    if st.button("Submit Entry"):
        total = cash_slips + online_slips
        commission = cash_slips * 30 + online_slips * 20

        new_data = {
            "Date": date.strftime("%Y-%m-%d"),
            "Branch": branch_name,
            "Rider": rider,
            "Cash Slips": cash_slips,
            "Online Slips": online_slips,
            "Total Slips": total,
            "Commission": commission
        }

        new_df = pd.DataFrame([new_data])

        with pd.ExcelWriter(DATA_FILE, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            book = writer.book
            if branch_name in book.sheetnames:
                existing_df = pd.read_excel(DATA_FILE, sheet_name=branch_name)
                updated_df = pd.concat([existing_df, new_df], ignore_index=True)
            else:
                updated_df = new_df

            updated_df.to_excel(writer, sheet_name=branch_name, index=False)

        st.success("✅ Slip submitted and saved to your branch sheet.")

else:
    if branch_code != "":
        st.error("❌ Invalid Branch Code. Please try again.")
